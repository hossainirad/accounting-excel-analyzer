# -*- coding: utf-8 -*-

from openpyxl.styles import Alignment
from openpyxl import load_workbook
from PyQt5 import QtCore, QtGui, QtWidgets

from peewee import *


from peewee import *


db = SqliteDatabase('acounting-db.db')


class CheckModel(Model):
    obj_id = CharField(unique=True, index=True)
    number = SmallIntegerField()
    amount = IntegerField()
    recieved_docs = CharField()
    condition = CharField()
    date_check = DateField()
    date_recieved_ckeck = DateField()
    bank_name = CharField(null=True)
    submit_date = DateField(null=True)

    class Meta:
        database = db # This model uses the "people.db" database.



def initial_db():
    db.connect()
    db.create_tables([CheckModel])
    db.close()



def str_to_date_converter(str_date):
    return str_date.replace('/', '-')


def open_excel(file):

    SEET_NAME = 'Sheet1'
    # file_name = input("please enter your file: ")

    wb = load_workbook(filename=file)
    # sheet
    ws = wb[SEET_NAME]
    _new = 0
    _duplicate = 0
    new_records = []
    for row in range(2, ws.max_row+1):
        # create id from number and date_check
        number = ws['A' + str(row)].value
        amount = ws['B' + str(row)].value
        recieved_docs = ws['E' + str(row)].value
        condition = ws['R' + str(row)].value
        date_check = str_to_date_converter(ws['S' + str(row)].value)
        date_recieved_ckeck = str_to_date_converter(ws['T' + str(row)].value)
        bank_name = ws['V' + str(row)].value
        obj_id = str(number)+date_recieved_ckeck

        exist_records = CheckModel.select(CheckModel.obj_id).where(CheckModel.obj_id == obj_id)



        if date_recieved_ckeck < '1401-01-01' and not len(exist_records):
            _new += 1
        elif condition in ["در جريان وصول", "برگشتى"] and not len(exist_records):
            submit_record_in_db([
                number,
                amount,
                recieved_docs,
                condition,
                date_check,
                date_recieved_ckeck,
                bank_name,
                None,
            ])
        elif not len(exist_records):
            # add new records to new_records
            new_records.append([
                    number,
                    amount,
                    recieved_docs,
                    condition,
                    date_check,
                    date_recieved_ckeck,
                    bank_name,
            ])
            _new += 1
        else:
            _duplicate += 1

    return new_records


def submit_record_in_db(record):
    condition = record[3]
    submit_date = record[7]
    if condition in ["در جريان وصول", "برگشتى"]:
        submit_date = "راکد"
    elif not submit_date:
        submit_date = "غیر بنفش"
    CheckModel.create(
        obj_id=str(record[0])+record[5],
        number=record[0],
        amount=record[1],
        recieved_docs=record[2],
        condition=record[3],
        date_check=record[4],
        date_recieved_ckeck=record[5],
        bank_name=record[6],
        submit_date=submit_date,
    )


def make_number_amount_comma_seperated(number):
    return ("{:,}".format(number))

def make_number_amount_comma_unseperated(num_string):
    return int(num_string.replace(',', ''))

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        # main window
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1000, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        # select file button
        self.select_file_btn = QtWidgets.QPushButton(self.centralwidget)
        self.select_file_btn.setGeometry(QtCore.QRect(450, 50, 91, 31))
        self.select_file_btn.setStyleSheet("font-size: 15px;")
        self.select_file_btn.setObjectName("pushButton")
        MainWindow.setCentralWidget(self.centralwidget)

        # submit record button
        self.submit_record_btn = QtWidgets.QPushButton(self.centralwidget)
        self.submit_record_btn.setGeometry(QtCore.QRect(300, 50, 91, 31))
        self.submit_record_btn.setStyleSheet("font-size: 15px;")
        self.submit_record_btn.setObjectName("submit_record_btn")
        self.submit_record_btn.setHidden(True)


        self.new_check_table_show = QtWidgets.QTableWidget(self.centralwidget)
        self.new_check_table_show.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.new_check_table_show.setGeometry(QtCore.QRect(0, 90, 1000, 571))
        self.new_check_table_show.setEditTriggers(QtWidgets.QAbstractItemView.DoubleClicked)
        self.new_check_table_show.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)
        self.new_check_table_show.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.new_check_table_show.horizontalHeader().setSortIndicatorShown(True)
        self.new_check_table_show.setObjectName("tableWidget")


        self.selected_rows = []

        font = QtGui.QFont()
        # font.setFamily("Times New Roman")
        font.setBold(True)
        font.setWeight(75)

        self.new_check_table_show.setColumnCount(8)
        # self.new_check_table_show.setRowCount(3)
        self.new_check_table_show.setHidden(True)
        for item_index in range(8):
            item = QtWidgets.QTableWidgetItem()
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            item.setFont(font)
            self.new_check_table_show.setHorizontalHeaderItem(item_index, item)


        self.splitter = QtWidgets.QSplitter(self.centralwidget)
        self.splitter.setGeometry(QtCore.QRect(280, 560, 451, 31))
        self.splitter.setOrientation(QtCore.Qt.Horizontal)
        self.splitter.setObjectName("splitter")
        self.splitter.setHidden(True)

        font = QtGui.QFont()
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)

        self.sum_amount_holder = QtWidgets.QPlainTextEdit(self.splitter)
        self.sum_amount_holder.setFont(font)
        self.sum_amount_holder.setReadOnly(True)
        self.sum_amount_holder.setObjectName("plainTextEdit")

        self.show_sum_line_edit = QtWidgets.QTextEdit(self.splitter)
        self.show_sum_line_edit.setFont(font)
        self.show_sum_line_edit.setReadOnly(True)
        self.show_sum_line_edit.setObjectName("textEdit")


        self.buttonPrint = QtWidgets.QPushButton(self.centralwidget)
        self.buttonPrint.setHidden(True)
        self.buttonPrint.setGeometry(QtCore.QRect(0, 300, 91, 31))
        self.buttonPrint.setStyleSheet("font-size: 15px;")
        self.buttonPrint.setObjectName("buttonPrint")
        # self.buttonPrint.setHidden(True)

        self.buttonPreview= QtWidgets.QPushButton(self.centralwidget)
        self.buttonPreview.setHidden(True)
        self.buttonPreview.setGeometry(QtCore.QRect(0, 360, 91, 31))
        self.buttonPreview.setStyleSheet("font-size: 15px;")
        self.buttonPreview.setObjectName("buttonPreview")
        # self.buttonPreview.setHidden(True)


        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        # signals and functions
        self.select_file_btn.clicked.connect(self.file_select)
        self.submit_record_btn.clicked.connect(self.submit_selected_record_in_db)
        self.buttonPrint.clicked.connect(self.handlePrint)
        self.buttonPreview.clicked.connect(self.handlePreview)

        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "گاز ورمه"))
        self.select_file_btn.setText(_translate("MainWindow", "انتخاب فایل"))
        self.buttonPrint.setText(_translate("MainWindow", "پرینت"))
        self.buttonPreview.setText(_translate("MainWindow", "پیشنمایش پرینت"))
        self.submit_record_btn.setText(_translate("MainWindow", "غیربنفش"))
        item = self.new_check_table_show.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "شماره"))
        item = self.new_check_table_show.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "مبلغ"))
        item = self.new_check_table_show.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "اسناد دریافتی"))
        item = self.new_check_table_show.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "وضعیت"))
        item = self.new_check_table_show.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "تاریخ چک"))
        item = self.new_check_table_show.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "تاریخ دریافت چک"))
        item = self.new_check_table_show.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "نام بانک"))
        item = self.new_check_table_show.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "تاریخ ثبت(اختیاری)"))

        self.show_sum_line_edit.setHtml(_translate("MainWindow",
                                                   "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
                                                   "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
                                                   "p, li { white-space: pre-wrap; }\n"
                                                   "</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:19pt; font-weight:600; font-style:normal;\">\n"
                                                   "<p align=\"center\" dir=\'rtl\' style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:11pt; font-weight:400;\">جمع مبالغ چک ها:</span></p></body></html>"))

    def file_select(self):
        self.file_select = QtWidgets.QFileDialog.getOpenFileName(MainWindow, 'Please give the excel file Mr. Kooti')
        sheet = open_excel(self.file_select[0])
        # remove_duplicate_records_from_excel()
        self.fill_table_items(sheet)

    def fill_table_items(self, list_item):
        # self.new_check_table_show.clear()
        self.new_check_table_show.setHidden(False)
        _translate = QtCore.QCoreApplication.translate
        # self.new_check_list_show = list_item
        self.new_check_table_show.setRowCount(len(list_item))
        for item_index in range(len(list_item)):  # 0 1 2 3 4 5 6 7
            ## add record to table
            font = QtGui.QFont()
            font.setPointSize(9)
            item = QtWidgets.QTableWidgetItem()
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            item.setText(str(item_index+1))
            self.new_check_table_show.setVerticalHeaderItem(item_index, item)
            for record_index in range(len(list_item[item_index])):
                item = QtWidgets.QTableWidgetItem()
                if record_index == 1:
                    item.setText(make_number_amount_comma_seperated(list_item[item_index][record_index]))
                else:
                    item.setText(str(list_item[item_index][record_index]))
                self.new_check_table_show.setItem(item_index, record_index, item)  # row, column, item


        self.submit_record_btn.setHidden(False)
        self.splitter.setHidden(False)
        self.fill_sum_amount_holder()

    def change_item_background_style(self):
        item = QtWidgets.QListWidgetItem()
        brush = QtGui.QBrush(QtGui.QColor(170, 180, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        for item in self.new_check_list_show.selectedItems():
            item.setBackground(brush)

    def submit_selected_record_in_db(self):
        selected_rows = self.new_check_table_show.selectionModel().selectedRows()
        selected_row_number = []
        for selected_record in selected_rows:
            selected_row_number.append(selected_record.row())

            single_record = []
            for cell in range(8):

                cell_text = self.new_check_table_show.item(selected_record.row(), cell)
                if cell_text:
                    single_record.append(cell_text.text())
                else:
                    single_record.append(None)
            submit_record_in_db(single_record)

            # remove selected rows
        counter = 0
        selected_row_number.sort()
        for rm in selected_row_number:
            self.new_check_table_show.removeRow(rm - counter)
            counter += 1
        self.fill_sum_amount_holder()

    def fill_sum_amount_holder(self):
        self.sum_amount = 0
        for row in range(self.new_check_table_show.rowCount()):
            amount = self.new_check_table_show.item(row, 1).text()
            amount = make_number_amount_comma_unseperated(amount)
            self.sum_amount += amount
        self.sum_amount_holder.setPlainText(make_number_amount_comma_seperated(self.sum_amount))

    def print_widget(self):
        dialog = QtPrintSupport.QPrintDialog()
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            self.handlePaintRequest(dialog.printer())

        # Create printer
        printer = QtPrintSupport.QPrinter()
        # Create painter
        painter = QtGui.QPainter()
        # Start painter
        painter.begin(printer)
        # Grab a widget you want to print
        screen = self.new_check_table_show.grab()
        # printer.setOrientation(QtPrintSupport.QPrinter.Landscape)
        # Draw grabbed pixmap
        painter.drawPixmap(10, 10, screen)

        # End painting
        painter.end()

    def handlePrint(self):
        dialog = QtPrintSupport.QPrintDialog()
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            self.handlePaintRequest(dialog.printer())

    def handlePreview(self):
        dialog = QtPrintSupport.QPrintPreviewDialog()
        dialog.paintRequested.connect(self.handlePaintRequest)

        dialog.exec_()

    def handlePreview(self):
        dialog = QtPrintSupport.QPrintPreviewDialog()
        dialog.paintRequested.connect(self.handlePaintRequest)

        dialog.exec_()

    def handlePaintRequest(self, printer):
        print("printer --> ", printer.orientation())
        printer.setOrientation(QtPrintSupport.QPrinter.Landscape)
        print("printer 2--> ", printer.orientation())

        document = self.makeTableDocument()
        document.print_(printer)

    def makeTableDocument(self):
        document = QtGui.QTextDocument()
        # document.setDocumentLayout(QtGui.QAbstractTextDocumentLayout(document))
        cursor = QtGui.QTextCursor(document)
        rows = self.new_check_table_show.rowCount()
        columns = self.new_check_table_show.columnCount()
        table = cursor.insertTable(rows + 1, columns)
        format = table.format()
        format.setHeaderRowCount(1)
        table.setFormat(format)
        format = cursor.blockCharFormat()
        format.setFontWeight(QtGui.QFont.Bold)
        format.setLayoutDirection(QtCore.Qt.RightToLeft)
        # headers font
        font = QtGui.QFont()
        # font.setFamily("Times New Roman")
        font.setBold(True)
        font.setWeight(75)

        for column in range(columns):
            cursor.setCharFormat(format)
            # cursor.setTextAlignment(QtCore.Qt.AlignCenter)
            # cursor.setFont(font)
            cursor.insertText(
                self.new_check_table_show.horizontalHeaderItem(column).text())
            cursor.movePosition(QtGui.QTextCursor.NextCell)

        # fill data
        for row in range(rows):
            for column in range(columns):
                if self.new_check_table_show.item(row, column):
                    cursor.insertText(self.new_check_table_show.item(row, column).text())
                else:
                    cursor.insertText("")
                cursor.movePosition(QtGui.QTextCursor.NextCell)
        return document



if __name__ == "__main__":
    import sys
    initial_db()
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

