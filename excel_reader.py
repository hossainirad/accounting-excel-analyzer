from pprint import pprint
import os
# from openpyxl import Workbook
from check_db import CheckModel
from openpyxl import load_workbook


def str_to_date_converter(str_date):
    return str_date.replace('/', '-')


def open_excel(file):

    SEET_NAME = 'Sheet1'
    # file_name = input("please enter your file: ")

    wb = load_workbook(filename=file)
    # sheet
    ws = wb[SEET_NAME]
    _new = 0
    _duplicate = 0
    new_records = []
    for row in range(2, ws.max_row+1):
        # create id from number and date_check
        number = ws['A' + str(row)].value
        amount = ws['B' + str(row)].value
        recieved_docs = ws['E' + str(row)].value
        condition = ws['R' + str(row)].value
        date_check = str_to_date_converter(ws['S' + str(row)].value)
        date_recieved_ckeck = str_to_date_converter(ws['T' + str(row)].value)
        bank_name = ws['V' + str(row)].value
        obj_id = str(number)+date_recieved_ckeck

        exist_records = CheckModel.select(CheckModel.obj_id).where(CheckModel.obj_id == obj_id)



        if date_recieved_ckeck < '1401-01-01' and not len(exist_records):
            _new += 1
        elif not len(exist_records):
            # add new records to new_records
            new_records.append([
                    number,
                    amount,
                    recieved_docs,
                    condition,
                    date_check,
                    date_recieved_ckeck,
                    bank_name,
            ])
            _new += 1
        else:
            _duplicate += 1

    return new_records


def submit_record_in_db(record):
    print(
        record
    )
    CheckModel.create(
        obj_id=str(record[0])+record[5],
        number=record[0],
        amount=record[1],
        recieved_docs=record[2],
        condition=record[3],
        date_check=record[4],
        date_recieved_ckeck=record[5],
        bank_name=record[6],
        submit_date=record[7],
    )


def make_number_amount_comma_seperated(number):
    return ("{:,}".format(number))

def make_number_amount_comma_unseperated(num_string):
    return int(num_string.replace(',', ''))
